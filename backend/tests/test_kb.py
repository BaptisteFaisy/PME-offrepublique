"""M2 client KB tests with focus on tenant isolation."""

from __future__ import annotations

import io
import json

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy.dialects import postgresql

from app.api.routes.kb import _extract_corpus_file_text, _parse_metadata_json
from app.config import Settings
from app.main import app
from app.models.kb import KbCorpusChunk, KbCorpusDocument
from app.schemas.kb import (
    CorpusChunkCreate,
    CorpusTextIn,
    InternalLibraryTextIn,
    KbClientCreate,
    KbStructuredProfileIn,
)
from app.services.embeddings import embed_texts
from app.services.kb import (
    client_retrieval_statement,
    ingest_client_corpus_text,
    ingest_internal_library_text,
    semantic_chunk_text,
)
from app.services.kb_import import (
    build_structured_profile_template,
    parse_structured_profile_import,
)

client = TestClient(app)


def test_openapi_exposes_kb_routes() -> None:
    schema = client.get("/openapi.json").json()
    assert "/kb/clients" in schema["paths"]
    assert "/kb/clients/{client_id}/structured" in schema["paths"]
    assert "/kb/structured/import-template" in schema["paths"]
    assert "/kb/clients/{client_id}/structured/import" in schema["paths"]
    assert "/kb/clients/{client_id}/corpus/documents" in schema["paths"]
    assert "/kb/clients/{client_id}/corpus/text" in schema["paths"]
    assert "/kb/clients/{client_id}/corpus/files" in schema["paths"]
    assert "/kb/clients/{client_id}/retrieve" in schema["paths"]
    assert "/kb/internal-library/documents" in schema["paths"]
    assert "/kb/internal-library/documents/{document_id}" in schema["paths"]
    assert "/kb/internal-library/text" in schema["paths"]
    assert "/kb/internal-library/files" in schema["paths"]


def test_siren_validation_normalizes_spaces() -> None:
    payload = KbClientCreate(name="Nettoyage pilote", siren="123 456 789")
    assert payload.siren == "123456789"


@pytest.mark.parametrize("siren", ["123", "abcdefghi", "1234567890"])
def test_siren_validation_rejects_invalid_values(siren: str) -> None:
    with pytest.raises(ValueError):
        KbClientCreate(name="Nettoyage pilote", siren=siren)


def test_structured_profile_keeps_max_three_financial_exercises() -> None:
    payload = {
        "financials": [
            {"fiscal_year": 2023, "revenue_eur": 1000},
            {"fiscal_year": 2022, "revenue_eur": 900},
            {"fiscal_year": 2021, "revenue_eur": 800},
        ]
    }
    assert len(KbStructuredProfileIn.model_validate(payload).financials) == 3

    with pytest.raises(ValueError):
        KbStructuredProfileIn.model_validate(
            {
                "financials": [
                    {"fiscal_year": 2023, "revenue_eur": 1000},
                    {"fiscal_year": 2022, "revenue_eur": 900},
                    {"fiscal_year": 2021, "revenue_eur": 800},
                    {"fiscal_year": 2020, "revenue_eur": 700},
                ]
            }
        )


def test_structured_json_import_accepts_complete_profile() -> None:
    payload = {
        "client": {"name": "Proprete pilote", "siren": "123456789"},
        "financials": [
            {"fiscal_year": 2025, "revenue_eur": 1500000},
            {"fiscal_year": 2024, "revenue_eur": 1400000},
        ],
        "staff_transfer_notes": [
            {
                "title": "Article 7",
                "commercial_argument": "Reprise du personnel maitrisee.",
                "assumptions": ["effectifs transmis par le sortant"],
            }
        ],
    }

    imported = parse_structured_profile_import(
        "profile.json",
        json.dumps(payload).encode("utf-8"),
    )

    assert imported.client is not None
    assert imported.client.siren == "123456789"
    assert len(imported.financials) == 2
    assert imported.staff_transfer_notes[0].convention_article == "Article 7"


def test_structured_xlsx_import_maps_cleaning_specific_sheets() -> None:
    workbook = Workbook()
    client_sheet = workbook.active
    client_sheet.title = "client"
    client_sheet.append(["name", "siren"])
    client_sheet.append(["Proprete pilote", "123 456 789"])

    financials = workbook.create_sheet("financials")
    financials.append(["fiscal_year", "revenue_eur"])
    financials.append([2025, 1500000])

    products = workbook.create_sheet("product_materials")
    products.append(["category", "name", "ecolabels", "use_case"])
    products.append(["produit", "Nettoyant sols", "Ecolabel Europe; Ecocert", "Sols durs"])
    products.append(["materiel", "Autolaveuse", "", "Grandes surfaces"])

    article_7 = workbook.create_sheet("staff_transfer_notes")
    article_7.append(["title", "commercial_argument", "assumptions"])
    article_7.append(["Article 7", "Continuite sociale et operationnelle.", "liste sortant"])

    plans = workbook.create_sheet("cleaning_plans")
    plans.append(["plan_key", "name", "site_type"])
    plans.append(["bureaux", "Plan bureaux", "tertiaire"])

    zones = workbook.create_sheet("cleaning_plan_zones")
    zones.append(["plan_key", "zone", "frequency", "operating_mode", "products", "materials"])
    zones.append(
        [
            "bureaux",
            "Accueil",
            "quotidien",
            "Balayage humide puis lavage.",
            "Nettoyant sols",
            "Chariot; frange microfibre",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    imported = parse_structured_profile_import("profile.xlsx", buffer.getvalue())

    assert imported.client is not None
    assert imported.client.siren == "123456789"
    assert imported.product_materials[0].ecolabels == ["Ecolabel Europe", "Ecocert"]
    assert imported.staff_transfer_notes[0].convention_article == "Article 7"
    assert imported.cleaning_plans[0].zones[0].zone == "Accueil"
    assert imported.cleaning_plans[0].zones[0].materials == [
        "Chariot",
        "frange microfibre",
    ]


def test_structured_xlsx_template_is_parseable() -> None:
    template = build_structured_profile_template()
    imported = parse_structured_profile_import("template.xlsx", template)

    assert imported.client is not None
    assert imported.client.siren == "123456789"
    assert imported.product_materials[0].category == "produit"
    assert imported.staff_transfer_notes[0].convention_article == "Article 7"
    assert imported.cleaning_plans[0].zones[0].zone == "Accueil"


def test_semantic_chunking_preserves_article_7_section_metadata() -> None:
    chunks = semantic_chunk_text(
        "ARTICLE 7 - REPRISE DU PERSONNEL\n\n"
        "La reprise du personnel est presentee comme un engagement central.\n\n"
        "PLAN DE NETTOYAGE\n\n"
        "Zones, frequences et modes operatoires sont decrits.",
        max_chars=120,
    )

    assert [chunk.chunk_index for chunk in chunks] == list(range(len(chunks)))
    assert any("REPRISE DU PERSONNEL" in chunk.metadata["section"] for chunk in chunks)
    assert all(chunk.metadata["content_hash"] for chunk in chunks)


def test_corpus_chunk_embedding_must_match_pgvector_dimension() -> None:
    CorpusChunkCreate(chunk_index=0, content="ok", embedding=[0.0] * 1536)
    with pytest.raises(ValueError):
        CorpusChunkCreate(chunk_index=0, content="bad", embedding=[0.0] * 32)


def test_deterministic_embedding_provider_is_stable_and_1536_dimensions() -> None:
    settings = Settings(embedding_provider="deterministic", embedding_dimensions=1536)
    first = embed_texts(["reprise du personnel article 7"], settings)[0]
    second = embed_texts(["reprise du personnel article 7"], settings)[0]

    assert len(first) == 1536
    assert first == second
    assert abs(sum(value * value for value in first) - 1.0) < 1e-9


def test_corpus_file_text_extraction_supports_txt_upload_metadata() -> None:
    text, metadata = _extract_corpus_file_text(
        "procedure.txt",
        b"ARTICLE 7 - REPRISE DU PERSONNEL\n\nPlan de nettoyage type.",
        Settings(),
    )

    assert "ARTICLE 7" in text
    assert metadata == {"page_count": 1, "extraction_warnings": []}


def test_metadata_json_must_be_an_object() -> None:
    assert _parse_metadata_json('{"type_prestation": "proprete"}') == {
        "type_prestation": "proprete"
    }
    with pytest.raises(HTTPException):
        _parse_metadata_json("[1, 2, 3]")


def test_client_corpus_chunks_have_schema_level_client_isolation() -> None:
    chunk_table = KbCorpusChunk.__table__
    document_table = KbCorpusDocument.__table__

    assert chunk_table.c.client_id.nullable is False
    assert document_table.c.client_id.nullable is False

    fk_column_sets = {
        tuple(element.parent.name for element in constraint.elements)
        for constraint in chunk_table.foreign_key_constraints
    }
    assert ("document_id", "client_id") in fk_column_sets


def test_retrieval_statement_filters_client_chunks_and_documents() -> None:
    stmt = client_retrieval_statement(
        client_id="client-a",
        query_embedding=[0.1] * 1536,
        service_type="proprete",
        include_internal_library=True,
        limit=5,
    )
    compiled = str(
        stmt.compile(
            dialect=postgresql.dialect(),
            compile_kwargs={"literal_binds": False},
        )
    )

    assert "kb_corpus_chunk.client_id = " in compiled
    assert "kb_corpus_document.client_id = " in compiled
    assert "kb_internal_library_chunk" in compiled
    assert "kb_internal_library_document.is_active IS true" in compiled
    assert "kb_corpus_document.client_id = kb_corpus_chunk.client_id" in compiled
    assert "kb_corpus_document.outcome" in compiled
    assert "kb_corpus_document.document_date" in compiled


class _FakeDb:
    def __init__(self) -> None:
        self.added = []

    def add(self, model) -> None:
        self.added.append(model)

    def add_all(self, models) -> None:
        self.added.extend(models)

    def flush(self) -> None:
        return None


def test_text_ingestion_forces_chunks_to_requested_client_id() -> None:
    settings = Settings(embedding_provider="deterministic", embedding_dimensions=1536)
    db = _FakeDb()
    payload = {
        "document": {
            "title": "Memoire technique gagne",
            "source_type": "memoire_technique",
            "service_type": "proprete",
            "document_date": "2025-01-15",
            "outcome": "gagne",
            "language": "fr",
            "metadata": {"marche": "bureaux"},
        },
        "text": "ARTICLE 7 - REPRISE DU PERSONNEL\n\n"
        "Argument commercial central.\n\n"
        "PLAN DE NETTOYAGE\n\n"
        "Accueil: quotidien, balayage humide.",
        "max_chars": 400,
    }

    document, chunks = ingest_client_corpus_text(
        db,
        client_id="client-a",
        payload=CorpusTextIn.model_validate(payload),
        settings=settings,
    )

    assert document.client_id == "client-a"
    assert chunks
    assert all(chunk.client_id == "client-a" for chunk in chunks)
    assert all(chunk.document_id == document.id for chunk in chunks)
    assert all(chunk.service_type == "proprete" for chunk in chunks)
    assert all(chunk.embedding and len(chunk.embedding) == 1536 for chunk in chunks)


def test_internal_library_text_ingestion_has_no_client_scope() -> None:
    settings = Settings(embedding_provider="deterministic", embedding_dimensions=1536)
    db = _FakeDb()
    payload = {
        "document": {
            "title": "Trame Article 7",
            "source_type": "trame_generique",
            "service_type": "proprete",
            "language": "fr",
            "metadata": {"origine": "interne"},
        },
        "text": "ARTICLE 7 - REPRISE DU PERSONNEL\n\n"
        "Trame generique sans donnees client.\n\n"
        "PLAN DE NETTOYAGE TYPE\n\n"
        "Zones, frequences et modes operatoires generiques.",
        "max_chars": 400,
    }

    document, chunks = ingest_internal_library_text(
        db,
        payload=InternalLibraryTextIn.model_validate(payload),
        settings=settings,
    )

    assert document.source_type == "trame_generique"
    assert document.metadata_json == {"origine": "interne"}
    assert chunks
    assert all(chunk.document_id == document.id for chunk in chunks)
    assert all(not hasattr(chunk, "client_id") for chunk in chunks)
    assert all(chunk.embedding and len(chunk.embedding) == 1536 for chunk in chunks)

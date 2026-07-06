"""Structured profile imports for the M2 client KB."""

from __future__ import annotations

import io
import json
from collections.abc import Iterable
from datetime import date, datetime
from typing import Any

from pydantic import ValidationError

from app.schemas.kb import (
    CleaningPlanIn,
    CleaningPlanZoneIn,
    KbStructuredProfileIn,
)

_SIMPLE_LIST_FIELDS = {
    "ecolabels",
    "assumptions",
    "products",
    "materials",
}
_JSON_FIELDS = {
    "details",
    "measurable_results",
    "habilitations",
}
_LIST_SHEETS = {
    "financials",
    "headcounts",
    "product_materials",
    "certifications",
    "insurances",
    "market_references",
    "supervisors",
    "qse_rse_policies",
    "staff_transfer_notes",
}
_TEMPLATE_SHEETS: dict[str, list[str]] = {
    "client": ["name", "siren", "kbis_s3_key", "kbis_issued_on", "notes"],
    "financials": ["fiscal_year", "revenue_eur"],
    "headcounts": [
        "label",
        "total_headcount",
        "supervisors_count",
        "operations_staff_count",
        "details",
    ],
    "product_materials": [
        "category",
        "name",
        "brand",
        "quantity",
        "use_case",
        "ecolabels",
        "technical_sheet_s3_key",
    ],
    "certifications": ["name", "issuer", "document_s3_key", "obtained_on", "expires_on"],
    "insurances": [
        "insurance_type",
        "provider",
        "policy_number",
        "coverage_summary",
        "document_s3_key",
        "expires_on",
    ],
    "market_references": [
        "reference_client",
        "object",
        "amount_eur",
        "duration_months",
        "assigned_headcount",
        "contact_name",
        "contact_email",
        "contact_phone",
        "service_type",
        "measurable_results",
    ],
    "supervisors": ["full_name", "role", "years_experience", "cv_s3_key", "habilitations"],
    "qse_rse_policies": ["policy_type", "title", "summary", "document_s3_key", "updated_on"],
    "staff_transfer_notes": [
        "title",
        "convention_article",
        "staff_count",
        "obligations_summary",
        "commercial_argument",
        "risk_notes",
        "assumptions",
    ],
    "cleaning_plans": ["plan_key", "name", "site_type", "description"],
    "cleaning_plan_zones": [
        "plan_key",
        "zone",
        "frequency",
        "operating_mode",
        "products",
        "materials",
    ],
}
_TEMPLATE_EXAMPLES: dict[str, list[Any]] = {
    "client": ["Proprete pilote", "123456789", "", "2026-01-01", "Pilote bureaux IDF"],
    "financials": [2025, 1500000],
    "headcounts": ["2025", 120, 8, 112, '{"agences": 2}'],
    "product_materials": [
        "produit",
        "Nettoyant sols",
        "Marque",
        "20 L",
        "Sols durs",
        "Ecolabel Europe; Ecocert",
        "",
    ],
    "certifications": ["ISO 9001", "AFNOR", "", "2024-01-01", "2027-01-01"],
    "insurances": ["RC Pro", "Assureur", "POL-123", "Responsabilite civile", "", "2027-01-01"],
    "market_references": [
        "Client reference",
        "Nettoyage bureaux",
        250000,
        36,
        12,
        "Contact",
        "contact@example.com",
        "+33100000000",
        "proprete bureaux",
        '[{"label": "taux satisfaction", "value": "98%"}]',
    ],
    "supervisors": ["Alice Martin", "Responsable secteur", 10, "", '[{"name": "SST"}]'],
    "qse_rse_policies": ["QSE_RSE", "Politique QSE/RSE", "Synthese", "", "2026-01-01"],
    "staff_transfer_notes": [
        "Article 7",
        "Article 7",
        12,
        "Reprise selon convention collective",
        "Continuite sociale et operationnelle.",
        "Verifier liste sortant",
        "liste sortant; anciennete",
    ],
    "cleaning_plans": ["bureaux", "Plan bureaux", "tertiaire", "Plan type bureaux"],
    "cleaning_plan_zones": [
        "bureaux",
        "Accueil",
        "quotidien",
        "Balayage humide puis lavage.",
        "Nettoyant sols",
        "Chariot; frange microfibre",
    ],
}


class StructuredImportError(ValueError):
    """Raised when a structured KB import file cannot be parsed."""


def build_structured_profile_template() -> bytes:
    """Build an XLSX template whose sheets match the structured import parser."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    workbook = openpyxl.Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["KB client - import structure"])
    readme.append(["Remplir les feuilles metier puis importer le fichier XLSX."])
    readme.append(["Ne pas renommer les feuilles ni les colonnes."])
    readme.append(["Listes simples: separer les valeurs par point-virgule."])
    readme.append(["Champs JSON: details, measurable_results, habilitations."])
    readme["A1"].font = Font(bold=True, size=14)
    readme.column_dimensions["A"].width = 78

    header_fill = PatternFill("solid", fgColor="D9E8FF")
    for sheet_name, headers in _TEMPLATE_SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        sheet.append(_TEMPLATE_EXAMPLES[sheet_name])
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[_column_letter(index)].width = max(14, len(header) + 4)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def parse_structured_profile_import(filename: str, data: bytes) -> KbStructuredProfileIn:
    """Parse a JSON or XLSX structured profile import into the API schema."""
    lower = filename.lower()
    try:
        if lower.endswith(".json"):
            payload = json.loads(data.decode("utf-8-sig"))
            return KbStructuredProfileIn.model_validate(payload)
        if lower.endswith(".xlsx"):
            return _parse_xlsx(data)
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise StructuredImportError(f"{filename}: JSON invalide") from exc
    except ValidationError as exc:
        raise StructuredImportError(f"{filename}: donnees invalides ({exc})") from exc

    raise StructuredImportError("Format import structure non supporte. Attendu: .json ou .xlsx")


def _parse_xlsx(data: bytes) -> KbStructuredProfileIn:
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise StructuredImportError(f"XLSX illisible ({exc})") from exc

    try:
        sheets = {_normalize_sheet_name(sheet.title): sheet for sheet in workbook.worksheets}
        payload: dict[str, Any] = {}

        client_rows = _sheet_rows(sheets.get("client"))
        if client_rows:
            payload["client"] = client_rows[0]

        for sheet_name in _LIST_SHEETS:
            rows = _sheet_rows(sheets.get(sheet_name))
            if rows:
                payload[sheet_name] = rows

        cleaning_plans = _cleaning_plans_from_sheets(
            _sheet_rows(sheets.get("cleaning_plans")),
            _sheet_rows(sheets.get("cleaning_plan_zones")),
        )
        if cleaning_plans:
            payload["cleaning_plans"] = cleaning_plans

        return KbStructuredProfileIn.model_validate(payload)
    finally:
        workbook.close()


def _cleaning_plans_from_sheets(
    plan_rows: list[dict[str, Any]],
    zone_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if not plan_rows:
        return []

    zones_by_key: dict[str, list[CleaningPlanZoneIn]] = {}
    for row in zone_rows:
        key = str(row.pop("plan_key", None) or row.pop("plan_name", None) or "").strip()
        if not key:
            raise StructuredImportError(
                "La feuille cleaning_plan_zones doit contenir plan_key ou plan_name"
            )
        zones_by_key.setdefault(key, []).append(CleaningPlanZoneIn.model_validate(row))

    plans: list[dict[str, Any]] = []
    for index, row in enumerate(plan_rows, start=1):
        key = str(row.pop("plan_key", None) or row.get("name") or f"plan-{index}").strip()
        plan = CleaningPlanIn.model_validate({**row, "zones": zones_by_key.get(key, [])})
        plans.append(plan.model_dump())
    return plans


def _sheet_rows(sheet) -> list[dict[str, Any]]:
    if sheet is None:
        return []

    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration:
        return []

    normalized_headers = [_normalize_header(value) for value in headers]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        row = _row_from_values(normalized_headers, values)
        if row:
            rows.append(row)
    return rows


def _row_from_values(headers: Iterable[str], values: Iterable[Any]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for header, value in zip(headers, values, strict=False):
        if not header or value is None or value == "":
            continue
        row[header] = _coerce_cell(header, value)
    return row


def _coerce_cell(header: str, value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if header in _SIMPLE_LIST_FIELDS:
        return _parse_list(value)
    if header in _JSON_FIELDS:
        return _parse_json_cell(header, value)
    return value


def _parse_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return value
    if not isinstance(value, str):
        return [str(value)]
    stripped = value.strip()
    if not stripped:
        return []
    if stripped.startswith("["):
        parsed = json.loads(stripped)
        if not isinstance(parsed, list):
            raise StructuredImportError("Une liste JSON etait attendue")
        return [str(item) for item in parsed]
    separator = ";" if ";" in stripped else ","
    return [part.strip() for part in stripped.split(separator) if part.strip()]


def _parse_json_cell(header: str, value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return value
    if not isinstance(value, str):
        raise StructuredImportError(f"{header}: JSON attendu")
    try:
        return json.loads(value)
    except json.JSONDecodeError as exc:
        raise StructuredImportError(f"{header}: JSON invalide") from exc


def _normalize_sheet_name(value: str) -> str:
    return value.strip().lower().replace(" ", "_").replace("-", "_")


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_").replace("-", "_")


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
